"""Python document extraction: PDF (PyMuPDF) and Excel (openpyxl/pandas).

Dual-verification architecture: this module is the *primary* extractor.
Gemini vision acts only as an independent verification layer on rendered
page images (see verification.py) — never as the sole source of numbers.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF

from ..financial.canonical import detect_statement_type, map_label_to_metric
from ..financial.numbers import parse_amount, detect_unit_multiplier

RENDER_DPI = 200

_PERIOD_RE = re.compile(
    # 20\d{2} must precede \d{2} in the alternation: Python alternation is ordered,
    # so the two-digit branch would match "20" of "2024" and turn 2023-2024 into
    # FY2023-20. The trailing (?!\d) stops "20240331" parsing as 2024-03.
    r"(?:FY\s?)?(20\d{2})\s?[-–—/]\s?(20\d{2}|\d{2})(?!\d)|march\s+(20\d{2})|31\.03\.(20\d{2})",
    re.IGNORECASE,
)


def sha256_of(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


@dataclass
class PageInfo:
    page_number: int  # 1-based
    statement_type: str
    score: int
    text_chars: int
    is_candidate: bool


@dataclass
class ExtractedItem:
    label: str
    metric: str | None
    period_label: str
    raw_value: str
    value: float | None
    page_number: int
    statement_type: str
    method: str = "pymupdf"


@dataclass
class PdfExtraction:
    page_count: int
    has_native_text: bool
    pages: list[PageInfo] = field(default_factory=list)
    items: list[ExtractedItem] = field(default_factory=list)
    unit_multiplier: float = 1.0
    unit_name: str = ""
    period_labels: list[str] = field(default_factory=list)


def normalise_period_label(text: str) -> str | None:
    """'2024-25', 'FY 2024-25', '2024-2025', 'March 2025' → 'FY2024-25'."""
    m = _PERIOD_RE.search(text)
    if not m:
        return None
    if m.group(1):
        start = int(m.group(1))
        end2 = int(m.group(2)) % 100
        # A fiscal year spans consecutive years. Rejecting anything else keeps a
        # bad guess ("FY2024-20", a date fragment) from becoming a phantom period
        # sitting alongside the periods actually read off the statement.
        if end2 != (start + 1) % 100:
            return None
        return f"FY{start}-{end2:02d}"
    year = int(m.group(3) or m.group(4))
    return f"FY{year - 1}-{year % 100:02d}"


def inspect_pdf(path: str | Path) -> PdfExtraction:
    """Open the PDF, classify pages, decide whether it has native text."""
    doc = fitz.open(str(path))
    pages: list[PageInfo] = []
    total_chars = 0
    for i, page in enumerate(doc):
        text = page.get_text("text")
        total_chars += len(text.strip())
        stype, score = detect_statement_type(text)
        pages.append(PageInfo(
            page_number=i + 1, statement_type=stype, score=score,
            text_chars=len(text.strip()),
            is_candidate=stype in ("balance_sheet", "profit_and_loss", "cash_flow") and score >= 3,
        ))
    has_text = total_chars > 200 * max(len(doc), 1) * 0.1  # heuristic
    result = PdfExtraction(page_count=len(doc), has_native_text=has_text, pages=pages)
    doc.close()
    return result


def _detect_document_unit(text: str) -> tuple[float, str]:
    """Look for '₹ in Lakhs' / 'Rs. in Crores' style headers."""
    for pattern, mult, name in [
        (r"(₹|rs\.?|rupees|amount)s?\s*(in)?\s*lakh", 100_000.0, "lakh"),
        (r"(₹|rs\.?|rupees|amount)s?\s*(in)?\s*crore", 10_000_000.0, "crore"),
        (r"(₹|rs\.?|rupees|amount)s?\s*(in)?\s*(thousand|'000)", 1_000.0, "thousand"),
        (r"(₹|rs\.?|rupees|amount)s?\s*(in)?\s*million", 1_000_000.0, "million"),
    ]:
        if re.search(pattern, text, re.IGNORECASE):
            return mult, name
    return 1.0, ""


_LINE_RE = re.compile(
    r"^(?P<label>[A-Za-z][A-Za-z&().,'/\- ]{2,60}?)\s{1,}"
    r"(?P<v1>\(?-?[\d,]+(?:\.\d+)?\)?)"
    r"(?:\s{1,}(?P<v2>\(?-?[\d,]+(?:\.\d+)?\)?))?\s*$"
)


def extract_pdf_items(path: str | Path, extraction: PdfExtraction | None = None) -> PdfExtraction:
    """Extract candidate line items from statement pages.

    Primary strategy: PyMuPDF's table detector (`page.find_tables()`), which
    reconstructs rows/columns from cell geometry — needed because many
    generated PDFs place each table cell as its own text object, so the
    naive "one physical text line per row" regex below never sees a label
    and its values on the same line (each ends up alone on its own line).
    Falls back to the line-based regex heuristic for pages with no
    detectable table (e.g. loosely formatted statements)."""
    ext = extraction or inspect_pdf(path)
    doc = fitz.open(str(path))

    # detect document-level unit and period columns from candidate pages
    header_text = ""
    for p in ext.pages:
        if p.is_candidate:
            header_text += doc[p.page_number - 1].get_text("text")[:2000] + "\n"
    if not header_text:
        header_text = doc[0].get_text("text")[:2000] if len(doc) else ""
    ext.unit_multiplier, ext.unit_name = _detect_document_unit(header_text)

    for p in ext.pages:
        if not p.is_candidate:
            continue
        page = doc[p.page_number - 1]

        table_items = _extract_table_items(page, p, ext)
        if table_items:
            ext.items.extend(table_items)
            continue

        _extract_line_items_regex(page, p, ext)
    doc.close()
    return ext


def _extract_table_items(page, p: PageInfo, ext: PdfExtraction) -> list[ExtractedItem]:
    """Reconstruct rows via PyMuPDF's geometry-based table detector."""
    items: list[ExtractedItem] = []
    try:
        tables = page.find_tables()
    except Exception:
        return items

    for table in tables.tables:
        rows = table.extract()
        if len(rows) < 2:
            continue
        header = rows[0]
        col_periods: dict[int, str] = {}
        for idx, cell in enumerate(header):
            if not cell:
                continue
            lbl = normalise_period_label(str(cell))
            if lbl:
                col_periods[idx] = lbl
        if not col_periods:
            continue  # no recognisable period columns — not a financial table

        for lbl in col_periods.values():
            if lbl not in ext.period_labels:
                ext.period_labels.append(lbl)

        for row in rows[1:]:
            if not row:
                continue
            label_cell = next((c for c in row if c and re.search(r"[A-Za-z]", str(c))), None)
            if label_cell is None:
                continue
            label = re.sub(r"\s+", " ", str(label_cell).strip()).rstrip(".").strip()
            if len(label) < 3 or label.lower().startswith(("note", "particular", "as at", "year ended")):
                continue
            metric = map_label_to_metric(label)
            for col_idx, period in col_periods.items():
                if col_idx >= len(row) or row[col_idx] is label_cell:
                    continue
                raw_val = row[col_idx]
                if raw_val is None:
                    continue
                value = parse_amount(raw_val, default_unit_multiplier=ext.unit_multiplier)
                if value is None:
                    continue
                items.append(ExtractedItem(
                    label=label, metric=metric, period_label=period,
                    raw_value=str(raw_val), value=value,
                    page_number=p.page_number, statement_type=p.statement_type,
                    method="pymupdf_table",
                ))
    return items


def _extract_line_items_regex(page, p: PageInfo, ext: PdfExtraction) -> None:
    """Fallback for pages with no detectable table: single-physical-line
    heuristic (label and value(s) whitespace-separated on one text line)."""
    text = page.get_text("text")

    # find period labels in page header (first ~15 lines)
    head_lines = text.splitlines()[:15]
    period_labels: list[str] = []
    for line in head_lines:
        for token in re.split(r"\s{2,}|\t", line):
            lbl = normalise_period_label(token)
            if lbl and lbl not in period_labels:
                period_labels.append(lbl)
    # statements list current year first
    if len(period_labels) >= 2:
        period_labels = period_labels[:2]
    for lbl in period_labels:
        if lbl not in ext.period_labels:
            ext.period_labels.append(lbl)

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        m = _LINE_RE.match(line.strip())
        if not m:
            continue
        label = m.group("label").strip().rstrip(".").strip()
        if len(label) < 3 or label.lower().startswith(("note", "particular", "as at", "year ended")):
            continue
        metric = map_label_to_metric(label)
        values = [m.group("v1"), m.group("v2")]
        for col, raw_val in enumerate(values):
            if raw_val is None:
                continue
            value = parse_amount(raw_val, default_unit_multiplier=ext.unit_multiplier)
            if value is None:
                continue
            period = (period_labels[col] if col < len(period_labels)
                      else (period_labels[0] if period_labels else ""))
            ext.items.append(ExtractedItem(
                label=label, metric=metric, period_label=period or "",
                raw_value=raw_val, value=value,
                page_number=p.page_number, statement_type=p.statement_type,
            ))


def render_pages(path: str | Path, page_numbers: list[int], out_dir: str | Path,
                 doc_id: str, dpi: int = RENDER_DPI) -> dict[int, str]:
    """Render selected pages to high-resolution PNGs; returns page → file path."""
    out: dict[int, str] = {}
    doc = fitz.open(str(path))
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    for pn in page_numbers:
        if pn < 1 or pn > len(doc):
            continue
        pix = doc[pn - 1].get_pixmap(matrix=mat)
        fname = str(Path(out_dir) / f"{doc_id}_p{pn}_{dpi}dpi.png")
        pix.save(fname)
        out[pn] = fname
    doc.close()
    return out


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def extract_xlsx_items(path: str | Path) -> PdfExtraction:
    """Extract line items from an Excel workbook of financial statements."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ext = PdfExtraction(page_count=len(wb.sheetnames), has_native_text=True)

    for sheet_idx, ws in enumerate(wb.worksheets):
        sheet_text = " ".join(
            str(c) for row in ws.iter_rows(max_row=10, values_only=True)
            for c in row if c is not None
        )
        stype, score = detect_statement_type(sheet_text + " " + ws.title)
        ext.pages.append(PageInfo(
            page_number=sheet_idx + 1, statement_type=stype, score=score,
            text_chars=len(sheet_text), is_candidate=stype != "other",
        ))
        unit_mult, unit_name = _detect_document_unit(sheet_text)
        if unit_mult != 1.0 and ext.unit_multiplier == 1.0:
            ext.unit_multiplier, ext.unit_name = unit_mult, unit_name

        # find period header row
        period_cols: dict[int, str] = {}
        for row in ws.iter_rows(max_row=12):
            for cell in row:
                if cell.value is None:
                    continue
                lbl = normalise_period_label(str(cell.value))
                if lbl and cell.column not in period_cols:
                    period_cols[cell.column] = lbl
            if len(period_cols) >= 2:
                break
        for lbl in period_cols.values():
            if lbl not in ext.period_labels:
                ext.period_labels.append(lbl)

        for row in ws.iter_rows(values_only=False):
            label_cell = next((c for c in row if isinstance(c.value, str) and len(c.value.strip()) > 2), None)
            if label_cell is None:
                continue
            label = label_cell.value.strip()
            metric = map_label_to_metric(label)
            for cell in row:
                if cell is label_cell or cell.value is None:
                    continue
                if not isinstance(cell.value, (int, float)):
                    continue
                period = period_cols.get(cell.column, "")
                ext.items.append(ExtractedItem(
                    label=label, metric=metric, period_label=period,
                    raw_value=str(cell.value),
                    value=float(cell.value) * ext.unit_multiplier,
                    page_number=sheet_idx + 1, statement_type=stype,
                    method="openpyxl",
                ))
    wb.close()
    return ext
